from flask import Flask, render_template, request, jsonify, send_file
import json
import os
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import requests
import base64
import time
import hashlib
import glob

app = Flask(__name__)
app.secret_key = 'clave_secreta_2024'

DB_FILE = 'participantes.db'
BACKUP_DIR = 'backups'

# Configuración simplificada - sin GitHub
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
REPO_OWNER = os.environ.get('REPO_OWNER', '')
REPO_NAME = os.environ.get('REPO_NAME', '')

def init_backup_dir():
    """Inicializar directorio de backups"""
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

def init_db():
    """Inicializar la base de datos local"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS participantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            telefono TEXT NOT NULL,
            genero TEXT NOT NULL,
            empresa TEXT,
            comentarios TEXT,
            fecha_inscripcion TEXT NOT NULL,
            timestamp TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def crear_backup_diario():
    """Crear backup diario acumulativo"""
    try:
        init_backup_dir()
        
        # Nombre del backup con fecha
        fecha_actual = datetime.now().strftime('%Y%m%d')
        backup_file = os.path.join(BACKUP_DIR, f'participantes_backup_{fecha_actual}.db')
        
        # Solo crear backup si no existe hoy o si la BD principal existe
        if not os.path.exists(backup_file) and os.path.exists(DB_FILE):
            import shutil
            shutil.copy2(DB_FILE, backup_file)
            print(f"✅ Backup diario creado: {backup_file}")
            
            # Limpiar backups antiguos (mantener últimos 7 días)
            limpiar_backups_antiguos()
            
            return True
        else:
            print("⏭️  Backup de hoy ya existe o no hay BD principal")
            return False
            
    except Exception as e:
        print(f"❌ Error creando backup diario: {e}")
        return False

def crear_backup_por_evento():
    """Crear backup por evento importante (nuevo registro, etc.)"""
    try:
        init_backup_dir()
        
        # Nombre del backup con fecha y hora exacta
        fecha_hora = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(BACKUP_DIR, f'participantes_evento_{fecha_hora}.db')
        
        if os.path.exists(DB_FILE):
            import shutil
            shutil.copy2(DB_FILE, backup_file)
            print(f"✅ Backup por evento creado: {backup_file}")
            return True
            
    except Exception as e:
        print(f"❌ Error creando backup por evento: {e}")
    return False

def limpiar_backups_antiguos(dias_retencion=7):
    """Eliminar backups más antiguos que los días de retención"""
    try:
        # Patrón para buscar archivos de backup
        patron = os.path.join(BACKUP_DIR, 'participantes_backup_*.db')
        archivos_backup = glob.glob(patron)
        
        fecha_limite = datetime.now() - timedelta(days=dias_retencion)
        
        for archivo in archivos_backup:
            # Extraer fecha del nombre del archivo
            nombre_archivo = os.path.basename(archivo)
            try:
                fecha_str = nombre_archivo.replace('participantes_backup_', '').replace('.db', '')
                fecha_backup = datetime.strptime(fecha_str, '%Y%m%d')
                
                if fecha_backup < fecha_limite:
                    os.remove(archivo)
                    print(f"🗑️  Backup antiguo eliminado: {archivo}")
                    
            except ValueError:
                # Si no puede parsear la fecha, mantener el archivo
                continue
                
    except Exception as e:
        print(f"❌ Error limpiando backups antiguos: {e}")

def obtener_info_backups():
    """Obtener información sobre los backups existentes"""
    try:
        init_backup_dir()
        patron = os.path.join(BACKUP_DIR, 'participantes_*.db')
        archivos_backup = glob.glob(patron)
        
        backups_info = []
        for archivo in archivos_backup:
            nombre = os.path.basename(archivo)
            tamaño = os.path.getsize(archivo)
            fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(archivo))
            
            backups_info.append({
                'nombre': nombre,
                'tamaño_bytes': tamaño,
                'tamaño_mb': round(tamaño / (1024 * 1024), 2),
                'fecha_modificacion': fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S'),
                'tipo': 'diario' if 'backup_' in nombre else 'evento'
            })
        
        # Ordenar por fecha (más reciente primero)
        backups_info.sort(key=lambda x: x['fecha_modificacion'], reverse=True)
        return backups_info
        
    except Exception as e:
        print(f"❌ Error obteniendo info de backups: {e}")
        return []

def get_db_connection():
    """Obtener conexión a la base de datos"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def deberia_crear_backup_diario():
    """Verificar si debería crear backup diario (una vez al día)"""
    try:
        init_backup_dir()
        
        # Buscar el backup más reciente
        patron = os.path.join(BACKUP_DIR, 'participantes_backup_*.db')
        archivos_backup = glob.glob(patron)
        
        if not archivos_backup:
            return True  # No hay backups, crear uno
            
        # Encontrar el backup más reciente
        archivos_backup.sort(reverse=True)
        backup_mas_reciente = archivos_backup[0]
        
        # Extraer fecha del nombre del archivo
        nombre_archivo = os.path.basename(backup_mas_reciente)
        fecha_str = nombre_archivo.replace('participantes_backup_', '').replace('.db', '')
        fecha_backup = datetime.strptime(fecha_str, '%Y%m%d')
        
        # Crear backup si el último es de ayer o anterior
        hoy = datetime.now().date()
        return fecha_backup.date() < hoy
        
    except Exception as e:
        print(f"❌ Error verificando backup diario: {e}")
        return True

# INICIALIZACIÓN CON BACKUP AUTOMÁTICO
print("🚀 Iniciando aplicación Flask...")
init_backup_dir()
init_db()

# Crear backup diario si es necesario
if deberia_crear_backup_diario():
    crear_backup_diario()
else:
    print("⏭️  Backup diario ya creado hoy")

print("✅ Aplicación lista con sistema de backups")

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/guardar', methods=['POST'])
def guardar_participante():
    try:
        data = request.json
        conn = get_db_connection()
        
        # Verificar duplicado
        existing = conn.execute(
            'SELECT id FROM participantes WHERE email = ?', 
            (data['email'].lower(),)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({'error': 'Email ya registrado'}), 400
        
        # Insertar nuevo participante
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        timestamp_actual = datetime.now().isoformat()
        
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO participantes 
            (nombre, email, telefono, genero, empresa, comentarios, fecha_inscripcion, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['nombre'],
            data['email'].lower(),
            data['telefono'],
            data['genero'],
            data.get('empresa', ''),
            data.get('comentarios', ''),
            fecha_actual,
            timestamp_actual
        ))
        
        participante_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Crear backup por evento importante
        crear_backup_por_evento()
        
        return jsonify({
            'success': True, 
            'message': 'Registro exitoso',
            'id': participante_id
        })
        
    except Exception as e:
        return jsonify({'error': f'Error del servidor: {str(e)}'}), 500

@app.route('/obtener', methods=['GET'])
def obtener_participantes():
    try:
        conn = get_db_connection()
        participantes = conn.execute('SELECT * FROM participantes ORDER BY id DESC').fetchall()
        conn.close()
        
        participantes_list = []
        for p in participantes:
            participantes_list.append({
                'id': p['id'],
                'nombre': p['nombre'],
                'email': p['email'],
                'telefono': p['telefono'],
                'genero': p['genero'],
                'empresa': p['empresa'] or '',
                'comentarios': p['comentarios'] or '',
                'fechaInscripcion': p['fecha_inscripcion'],
                'timestamp': p['timestamp']
            })
        
        return jsonify(participantes_list)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generar-excel', methods=['GET'])
def generar_excel():
    try:
        conn = get_db_connection()
        participantes = conn.execute('SELECT * FROM participantes ORDER BY id DESC').fetchall()
        conn.close()
        
        if not participantes:
            return jsonify({'error': 'No hay datos para exportar'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Participantes"
        
        headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'Género', 'Empresa', 'Comentarios', 'Fecha de Inscripción']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, participante in enumerate(participantes, 2):
            ws.cell(row=row, column=1, value=participante['id'])
            ws.cell(row=row, column=2, value=participante['nombre'])
            ws.cell(row=row, column=3, value=participante['email'])
            ws.cell(row=row, column=4, value=participante['telefono'])
            ws.cell(row=row, column=5, value=participante['genero'])
            ws.cell(row=row, column=6, value=participante['empresa'] or '')
            ws.cell(row=row, column=7, value=participante['comentarios'] or '')
            ws.cell(row=row, column=8, value=participante['fecha_inscripcion'])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'participantes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/eliminar-todos', methods=['POST'])
def eliminar_todos():
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM participantes')
        conn.commit()
        conn.close()
        
        # Crear backup antes de eliminar
        crear_backup_por_evento()
        
        return jsonify({'success': True, 'message': 'Todos los datos eliminados'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/estado', methods=['GET'])
def estado_db():
    try:
        conn = get_db_connection()
        count = conn.execute('SELECT COUNT(*) as total FROM participantes').fetchone()['total']
        conn.close()
        
        backups_info = obtener_info_backups()
        
        return jsonify({
            'estado': 'ok',
            'total_participantes': count,
            'bd_existe': os.path.exists(DB_FILE),
            'total_backups': len(backups_info),
            'backups': backups_info,
            'modo': 'local con backups automáticos'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/descargar-backup', methods=['GET'])
def descargar_backup():
    """Descargar la base de datos principal como backup"""
    try:
        if not os.path.exists(DB_FILE):
            return jsonify({'error': 'No existe la base de datos'}), 404
            
        return send_file(
            DB_FILE,
            as_attachment=True,
            download_name=f'backup_principal_{datetime.now().strftime("%Y%m%d_%H%M")}.db'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/descargar-backup/<nombre_backup>', methods=['GET'])
def descargar_backup_especifico(nombre_backup):
    """Descargar un backup específico"""
    try:
        backup_path = os.path.join(BACKUP_DIR, nombre_backup)
        if not os.path.exists(backup_path):
            return jsonify({'error': 'Backup no encontrado'}), 404
            
        return send_file(
            backup_path,
            as_attachment=True,
            download_name=nombre_backup
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/forzar-backup-diario', methods=['POST'])
def forzar_backup_diario():
    """Forzar la creación de un backup diario"""
    try:
        success = crear_backup_diario()
        if success:
            return jsonify({'success': True, 'message': 'Backup diario creado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'No se pudo crear el backup diario'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/limpiar-backups', methods=['POST'])
def limpiar_backups():
    """Limpiar backups antiguos"""
    try:
        data = request.json
        dias = data.get('dias', 7) if data else 7
        
        limpiar_backups_antiguos(dias)
        
        return jsonify({'success': True, 'message': f'Backups antiguos (más de {dias} días) eliminados'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
